import asyncio
import logging
import io
import datetime
from aiogram import Bot, Dispatcher, F, Router
from aiogram.filters import CommandStart, Command
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardMarkup, KeyboardButton, BufferedInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters.callback_data import CallbackData
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import database as db

# --- SOZLAMALAR ---
BOT_TOKEN = "BOT_TOKEN_SHU_YERGA_YOZING"
ADMIN_ID = 123456789 # O'z ID raqamingizni yozing
WORKERS_PER_PAGE = 5

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
router = Router()

# --- CALLBACK DATA FACTORY (Clean Code yondashuvi) ---
class WorkerCB(CallbackData, prefix="wrk"):
    action: str  # toggle, save, prev, next
    id: int = 0
    page: int = 0

class AdminCB(CallbackData, prefix="adm"):
    action: str
    user_id: int

# --- FSM HOLATLAR ---
class AddService(StatesGroup):
    car_model = State()
    modifications = State()
    price = State()
    coworkers = State()
    vin_code = State()

# --- KEYBOARDS ---
def main_menu():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="🚗 Yangi avtomobil qo'shish")], [KeyboardButton(text="📊 Mening natijalarim")]],
        resize_keyboard=True
    )

def cancel_keyboard():
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="⬅️ Orqaga"), KeyboardButton(text="❌ Bekor qilish")]],
        resize_keyboard=True
    )

# --- START & RO'YXATDAN O'TISH ---
@router.message(CommandStart())
async def start_cmd(message: Message):
    user = await db.get_user(message.from_user.id)
    if not user:
        await db.add_user(message.from_user.id, message.from_user.full_name)
        ikb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Tasdiqlash", callback_data=AdminCB(action="approve", user_id=message.from_user.id).pack()),
             InlineKeyboardButton(text="❌ Rad etish", callback_data=AdminCB(action="reject", user_id=message.from_user.id).pack())]
        ])
        await message.bot.send_message(ADMIN_ID, f"🆕 Yangi xodim:\nIsmi: {message.from_user.full_name}\nID: {message.from_user.id}", reply_markup=ikb)
        await message.answer("So'rov adminga yuborildi. Tasdiqlanishini kuting⏳")
    elif user['status'] == 'pending':
        await message.answer("So'rovingiz tasdiqlanmoqda⏳")
    elif user['status'] in ['approved', 'admin']:
        await message.answer("Asosiy menyuga xush kelibsiz!", reply_markup=main_menu())
    else:
        await message.answer("Siz tizimdan chetlashtirilgansiz🚫")

# --- ADMIN PANEL ---
@router.callback_query(AdminCB.filter(F.action == "approve"))
async def approve_user(call: CallbackQuery, callback_data: AdminCB):
    await db.update_user_status(callback_data.user_id, 'approved')
    await call.message.edit_text(f"{call.message.text}\n\n✅ Tasdiqlandi!")
    try: await call.bot.send_message(callback_data.user_id, "Tabriklaymiz! Admin sizni tasdiqladi🎉", reply_markup=main_menu())
    except: pass

@router.callback_query(AdminCB.filter(F.action == "reject"))
async def reject_user(call: CallbackQuery, callback_data: AdminCB):
    await db.update_user_status(callback_data.user_id, 'fired')
    await call.message.edit_text(f"{call.message.text}\n\n❌ Rad etildi!")

@router.message(Command("admin"))
async def admin_panel(message: Message):
    if message.from_user.id != ADMIN_ID: return
    total_cars, total_revenue = await db.get_general_stats()
    top_workers = await db.get_top_workers()
    
    text = f"👑 <b>Super Admin Panel</b>\n\n🚗 Mashinalar: <b>{total_cars} ta</b>\n💰 Umumiy tushum: <b>{total_revenue:,} UZS</b>\n\n🏆 <b>Top Xodimlar:</b>\n"
    for i, w in enumerate(top_workers, 1):
        text += f"{i}. {w[0]} - {w[1]} ish ({w[2]:,} UZS)\n"
    await message.answer(text, parse_mode="HTML")

# --- FSM: XIZMAT QO'SHISH ---
@router.message(F.text == "❌ Bekor qilish")
async def cancel_fsm(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Jarayon bekor qilindi.", reply_markup=main_menu())

@router.message(F.text == "🚗 Yangi avtomobil qo'shish")
async def start_add_service(message: Message, state: FSMContext):
    user = await db.get_user(message.from_user.id)
    if not user or user['status'] not in ['approved', 'admin']: return
    await state.set_state(AddService.car_model)
    await message.answer("Avto modeli va yili (Masalan: Labo, 2023):", reply_markup=cancel_keyboard())

@router.message(AddService.car_model)
async def process_car(message: Message, state: FSMContext):
    if message.text == "⬅️ Orqaga": return
    await state.update_data(car_model=message.text)
    await state.set_state(AddService.modifications)
    await message.answer("Qilingan ishlar (Masalan: Potolok fanera, chexol, pol rezinka):")

@router.message(AddService.modifications)
async def process_mods(message: Message, state: FSMContext):
    if message.text == "⬅️ Orqaga":
        await state.set_state(AddService.car_model)
        return await message.answer("Avto modelini qaytadan kiriting:")
    await state.update_data(modifications=message.text)
    await state.set_state(AddService.price)
    await message.answer("Summa (Faqat raqam, masalan: 1500000):")

@router.message(AddService.price)
async def process_price(message: Message, state: FSMContext):
    if message.text == "⬅️ Orqaga":
        await state.set_state(AddService.modifications)
        return await message.answer("Qilingan ishlarni qayta kiriting:")
    if not message.text.isdigit():
        return await message.answer("⚠️ Faqat raqam kiriting!")
    
    await state.update_data(price=int(message.text), selected_workers=[])
    await state.set_state(AddService.coworkers)
    await send_workers_keyboard(message, 0, [])

# --- PAGINATSIYALI INLINE KEYBOARD (Xodimlar uchun) ---
async def send_workers_keyboard(message: Message, page: int, selected: list, is_edit=False):
    total_workers = await db.count_active_workers()
    workers = await db.get_workers(limit=WORKERS_PER_PAGE, offset=page * WORKERS_PER_PAGE)
    
    kb = []
    for w in workers:
        status = "✅ " if w['id'] in selected else ""
        kb.append([InlineKeyboardButton(
            text=f"{status}{w['full_name']}", 
            callback_data=WorkerCB(action="toggle", id=w['id'], page=page).pack()
        )])
    
    # Navigatsiya tugmalari
    nav_buttons = []
    if page > 0:
        nav_buttons.append(InlineKeyboardButton(text="⬅️ Oldingi", callback_data=WorkerCB(action="prev", page=page-1).pack()))
    if (page + 1) * WORKERS_PER_PAGE < total_workers:
        nav_buttons.append(InlineKeyboardButton(text="Keyingi ➡️", callback_data=WorkerCB(action="next", page=page+1).pack()))
    if nav_buttons: kb.append(nav_buttons)
    
    kb.append([InlineKeyboardButton(text="💾 Saqlash", callback_data=WorkerCB(action="save", page=page).pack())])
    markup = InlineKeyboardMarkup(inline_keyboard=kb)
    
    text = "Kimlar bilan ishladingiz? O'zingizni ham belgilang:"
    if is_edit: await message.edit_text(text, reply_markup=markup)
    else: await message.answer(text, reply_markup=markup)

@router.callback_query(WorkerCB.filter())
async def process_worker_cb(call: CallbackQuery, callback_data: WorkerCB, state: FSMContext):
    data = await state.get_data()
    selected = data.get('selected_workers', [])
    
    if callback_data.action == "toggle":
        if callback_data.id in selected: selected.remove(callback_data.id)
        else: selected.append(callback_data.id)
        await state.update_data(selected_workers=selected)
        await send_workers_keyboard(call.message, callback_data.page, selected, is_edit=True)
        
    elif callback_data.action in ["prev", "next"]:
        await send_workers_keyboard(call.message, callback_data.page, selected, is_edit=True)
        
    elif callback_data.action == "save":
        if not selected: return await call.answer("Kamida 1 kishini tanlang!", show_alert=True)
        await state.set_state(AddService.vin_code)
        await call.message.delete()
        await call.message.answer("Ajoyib! Avtomobil raqamini yoki VIN kodini kiriting (Masalan: 01A777AA):")

# --- YAKUNIY BOSQICH VA EXCEL ---
@router.message(AddService.vin_code)
async def process_vin(message: Message, state: FSMContext):
    if message.text == "⬅️ Orqaga": return
    
    data = await state.get_data()
    vin_code = message.text.replace(" ", "_").upper()
    
    worker_names = await db.get_worker_names_by_ids(data['selected_workers'])
    data['worker_names'] = worker_names # Ismlarni Excel uchun saqlaymiz
    
    await db.add_service(data['car_model'], data['modifications'], data['price'], vin_code, data['selected_workers'])
    excel_file = await asyncio.to_thread(generate_premium_excel, data, vin_code)
    
    await message.answer_document(
        document=BufferedInputFile(excel_file, filename=f"Hisobot_{vin_code}.xlsx"),
        caption="✅ Muvaffaqiyatli saqlandi!",
        reply_markup=main_menu()
    )
    await state.clear()

# --- PREMIUM EXCEL GENERATOR (LIDER AVTOTEX 555 Uslubida) ---
def generate_premium_excel(data, vin_code):
    wb = Workbook()
    ws = wb.active
    ws.title = "Xizmat Hisoboti"
    
    # Asosiy sarlavha ("LIDER AVTOTEX 555")
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "LIDER AVTOTEX 555 - XIZMAT KO'RSATISH HISOBOTI"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    headers = ["Sana", "Avtomobil / VIN", "Bajarilgan ishlar", "Jami Summa", "Xodimlar"]
    ws.append(headers)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    today = datetime.datetime.now().strftime("%d.%m.%Y")
    ws.append([
        today,
        f"{data['car_model']} ({vin_code})",
        data['modifications'],
        f"{data['price']:,} UZS",
        ", ".join(data['worker_names'])
    ])
    
    # Chiziqlar va Avto-kenglik
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()

# --- BACKGROUND TASK ---
async def background_tasks():
    while True:
        await db.cleanup_old_records()
        await asyncio.sleep(86400)

async def main():
    await db.init_db()
    bot = Bot(token=BOT_TOKEN)
    dp = Dispatcher()
    dp.include_router(router)
    asyncio.create_task(background_tasks())
    logging.info("Premium Bot ishga tushdi...")
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == "__main__":
    asyncio.run(main())
